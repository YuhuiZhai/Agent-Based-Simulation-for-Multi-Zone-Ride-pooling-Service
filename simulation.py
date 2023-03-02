import numpy as np
from city import City
from fleet_taxi import Taxifleet
from eventQueue import EventQueue
from tqdm import tqdm
from animation import Animation
import xlsxwriter as xw
import openpyxl as op
import os
class Simulation:
    def __init__(self, city:City, dt:float, T:float, lmd_matrix=None, fleet_matrix=None, rebalance_matrix=None, 
                test_deliver=False, turning_random=True, rebalance_t_tp=None):
        self.clock, self.city, self.dt, self.T, self.turninng_random = 0, city, dt, T, turning_random
        self.city.assignT(T)
        self.lmd_m, self.fleet_m, self.rebalance_m = lmd_matrix, fleet_matrix, rebalance_matrix
        self.fleet = Taxifleet(self.fleet_m, 0, self.city, self.dt, self.T, self.rebalance_m, turning_random)
        self.events = EventQueue(T, self.lmd_m, 0, self.city)
        self.fleet_info = [] # list of [ax, ay], [sx, sy], [ix, iy]
        self.passenger_info = [] # list of [px, py]
        self.zone_number = self.city.n**2 
        self.folder_name = None
        # 3 dimensional list storing the traveling time of passenger 
        self.traveling_ts = [[[[], []] for j in range(self.zone_number)] for i in range(self.zone_number)]
        # double dictionary for state number and transition rate, just like fleet.zone_group
        self.status_number = {i:{} for i in range(self.zone_number)}
        # index of iteration
        self.idx = 0
        self.Ps = []
        # record of flow transition, representing [a, b, c, d, g, p]
        self.flow_table = [{}, {}, {}, {}, {}, {}]
        
        # dictionary storing unserved number of passenger
        self.unserved_number = {i:0 for i in range(self.zone_number)}
        # record of unserved information [info = [starting time, ozone, dzone, direction, ni0i0, sum_ni0j0]]
        self.unserved_record = []

        # indicator of testing deliver distance -> One turn if true
        self.test_deliver = test_deliver
        self.bound_m = None
        # dynamic rebalance matrix cumulation
        self.rebalance_m_tp = []
        self.rebalance_m_bound = []
        self.b = []
        self.rebalance_t_tp = rebalance_t_tp
        
    def name(self, folder_name):
        self.folder_name = folder_name
        return 
    
    def move(self):
        self.clock += self.dt
        self.fleet.fleet_move(self.dt, self.test_deliver)
        self.events.move(self.dt)
    
    def simple_serve(self):
        self.timeline = np.arange(0, self.T, self.dt)
        idx = 0
        if self.rebalance_t_tp != None:
            self.tp_N = int(len(self.timeline)/int((self.T/self.rebalance_t_tp)))
        for t in tqdm(self.timeline, desc="simple_serve loading"):
            p_id, p = self.events.head()
            while (not self.events.empty() and p.t_start < t):
                # unserved passenger exists
                opt_veh, dist, prev_status = self.fleet.serve(p)
                if opt_veh == None:
                    self.update_unserved(p)
                self.update_flow_table(0, prev_status)
                self.events.dequeue()
                p_id, p = self.events.head()
            if self.rebalance_t_tp == None:
                self.fleet.rebalance()
                # rebalance_m_bound = self.fleet.rebalance_by_bound(self.bound_m)
                # self.rebalance_m_bound.append(rebalance_m_bound)

            elif idx % self.tp_N == 0:
                fleet_m, b = self.fleet.rebalance_tp()
                self.rebalance_m_tp.append(fleet_m)
                self.b.append(b)
            self.move()
            self.update()
            idx += 1
        return 

    def add_bound(self, bound_m):
        self.bound_m = bound_m
        return 

    def test_simple_taxi(self):
        test_sample = 0
        count = 0
        for i in self.fleet.vehicles:
            if count > 3:
                break
            test_sample = self.fleet.vehicles[i]
            print(test_sample.status_record)
            count += 1
        return 

    def update_unserved(self, passenger):
        self.unserved_number[passenger.zone.id] += 1 
        # record is a list [starting time, ozone, dzone, direction, ni0j0 - ni0i0 - ni0j0]
        record = [0 for i in range(4+self.zone_number)]
        record[0] = passenger.t_start
        O, D, = passenger.odzone()
        record[1], record[2] = O+1, D+1
        if O == D: record[3] = self.city.direction_helper(*passenger.location())
        for j in range(self.zone_number):
            ni0j0 = self.status_number[O][(-1, j, -1)][-1] if (-1, j, -1) in self.status_number[O] else 0
            record[j+4] = ni0j0
        self.unserved_record.append(record)
        return record

    def update_state_transition(self):
        for zone_id in self.fleet.zone_group:
            for status in self.fleet.zone_group[zone_id]:
                if status not in self.status_number[zone_id]:
                    self.status_number[zone_id][status] = [0 for i in range(self.idx)]
                self.status_number[zone_id][status].append(len(self.fleet.zone_group[zone_id][status]))
        return
    
    def update_flow_table(self, type:int, status):
        if status == None:
            return 
        if status not in self.flow_table[type]:
            self.flow_table[type][status] = [0 for i in range(self.idx)]
        while len(self.flow_table[type][status]) < self.idx:
            self.flow_table[type][status].append(0)
        if len(self.flow_table) == self.idx:
            self.flow_table[type][status].append(1)
        else: 
            self.flow_table[type][status][-1] += 1
        return 

    def update_flow(self):
        # 0:a, 1:b, 2:c, 3:d, 4:g, 5:p
        for veh_id in self.fleet.vehicles:
            veh = self.fleet.vehicles[veh_id]
            prev_status = veh.prev_status
            curr_status = veh.status
            if prev_status == None or prev_status == curr_status:
                continue
            ps0, ps1, ps2, ps3 = prev_status
            cs0, cs1, cs2, cs3 = curr_status
            # status (j, i, -1, -1)
            if ps0 != ps1 and ps1 != -1 and ps2 == -1 and ps3 == -1:
                self.update_flow_table(1, prev_status)

            # status (i, -1, -1, -1) *
            elif ps1 == -1 and ps2 == -1 and ps3 == -1:
                # bij00 *
                if cs0 != cs1:
                    self.update_flow_table(1, curr_status)
                # ai000 *
                else:
                    self.update_flow_table(0, prev_status)

            # status (i, -1, i, -1) *
            elif ps1 == -1 and ps2 == ps0 and ps3 == -1:
                # di0i0 *
                if cs2 == -1:
                    self.update_flow_table(3, prev_status)
                # ai0i0 *
                else:
                    self.update_flow_table(0, prev_status)
            # status (i, -1, j, -1) * 
            elif ps1 == -1 and ps2 != -1 and ps2 != ps0 and ps3 == -1:
                # ai0j0 *
                if cs0 == cs1: 
                    self.update_flow_table(0, prev_status)
                # gi0j0 *
                if cs0 != ps0 and cs2 == ps2: 
                    self.update_flow_table(4, prev_status)
                    self.update_flow_table(2, curr_status)

            # status (i, i, -1 or i or j, -1) *
            elif ps0 == ps1 and ps3 == -1: 
                # piij0_j *
                if cs2 == ps2:
                    self.update_flow_table(5, (*prev_status, cs3))
                else:
                    self.update_flow_table(5, (*prev_status, cs2))
            
            # status (i, -1, i, k) *
            elif ps0 == ps2 and ps1 == -1 and ps3 != -1:
                self.update_flow_table(3, prev_status)
            
            # status (i, -1, j, k) *
            elif ps1 == -1 and ps0 != ps2 and ps2 != -1 and ps3 != -1:
                # gi0jk and ci0jk *
                self.update_flow_table(4, prev_status)
                self.update_flow_table(2, curr_status)
        return 
    
    def update_P(self):
        P = self.fleet.computeP()
        self.Ps.append(P)

    def update(self):
        ax, ay = [], []
        sx, sy = [], []
        ix, iy = [], []
        interx, intery = [], []
        px, py = [], []

        [px, py] = self.events.sketch_helper()
        px += px; py += py
        [ax, ay], [sx, sy], [ix, iy], [interx, intery] = self.fleet.sketch_helper()
        ax += ax; ay += ay; sx += sx; sy += sy
        ix += ix; iy += iy; interx += interx; intery += intery; 
        self.fleet_info.append([[ax, ay], [sx, sy], [ix, iy], [interx, intery]])
        self.passenger_info.append([[px, py]])
        self.update_state_transition()
        self.update_flow()
        self.update_P()
        self.idx += 1
        return 

    def make_animation(self, compression=50, fps=15, name="simulation", path=""):
        print("animation plotting")
        animation = Animation(self.city, self.fleet_info, self.passenger_info)
        fleet_pattern = ({0:"idle", 1:"assigned", 2:"in service", 3:"interchanged"},
                            {0:'g', 1:'y', 2:'r', 3:'k'},
                            'o'    
                        )
        passenger_pattern = ({0:"Passenger"}, 
                                {0:'b'},
                                'v'
                            )
        if name == "":
            animation.plot(compression, fps, fleet_pattern, passenger_pattern, "simulation", path)
        else: animation.plot(compression, fps, fleet_pattern, passenger_pattern, name, path)
        return 

    def export_flow(self, full=False):
        temp = {0:"a", 1:"b", 2:"c", 3:"d", 4:"g", 5:"p"}
        workbook = xw.Workbook(f"{self.fleet_m}_flow.xlsx")
        worksheet0 = workbook.add_worksheet("summary")
        worksheets = []
        for i in range(6):
            worksheet0.write(0, i*2, f"{temp[i]}_xxxx")
            if full:
                worksheet = workbook.add_worksheet(temp[i])
                worksheet.write(0, 0, "time")
                for j in range(self.idx):
                    worksheet.write(j+1, 0, j*self.dt)
            flows = self.flow_table[i]
            for j, status in enumerate(flows):
                worksheet0.write(j+1, i*2, f"{temp[i]}_{status}")
                worksheet0.write(j+1, i*2+1, sum(flows[status])/self.idx/self.dt)
                if full:
                    worksheet.write(0, j+1, f"{temp[i], status}")
                    for k in range(self.idx):
                        if k >= len(flows[status]):
                            worksheet.write(k+1, j+1, 0)
                        else:
                            worksheet.write(k+1, j+1, flows[status][k])
        workbook.close()
        return
    
    def export_P_over_lambda(self):
        Ps = self.Ps[int(len(self.Ps)*2/3):]
        avg_P = sum(Ps) / len(Ps)
        total_lambda = sum(np.array(self.lmd_m).flatten())
        print(avg_P / total_lambda)
        return avg_P / total_lambda
    
    # export state transition information 
    def export_state_number(self, full=False, shift=1):
        dir_path = os.path.dirname(os.path.realpath(__file__))
        path = f"{dir_path}\{self.folder_name}\state_number" if self.folder_name != None else f"{dir_path}\state_number"
        path_exist = os.path.exists(f"{path}")
        if not path_exist:
            os.makedirs(f"{path}")
        workbook = xw.Workbook(f"{path}\{self.fleet_m}_state_number.xlsx")
        cell_format = workbook.add_format()
        cell_format.set_bold()
        cell_format.set_font_color('red')
        worksheet0 = workbook.add_worksheet("summary")
        worksheet0.write(0, 0, "status")
        worksheet0.write(0, 1, "avg number")
        if full:
            worksheet1 = workbook.add_worksheet("status number")
        worksheet2 = workbook.add_worksheet("unserved number")
        row = 1
        ni000 = [0 for i in range(self.zone_number)]
        for zone_id in self.status_number:
            for status in self.status_number[zone_id]:
                l = len(self.status_number[zone_id][status])
                # get rid of warming stage
                avg = sum(self.status_number[zone_id][status][int(1/3*l):]) / len(self.status_number[zone_id][status][int(1/3*l):]) 
                if status == (-1, -1, -1):
                    worksheet0.write(row, 0, f"n_{zone_id+shift}{status[0]+shift}{status[1]+shift}{status[2]+shift}", cell_format)
                    ni000[zone_id] = avg
                else:
                    worksheet0.write(row, 0, f"n_{zone_id+shift}{status[0]+shift}{status[1]+shift}{status[2]+shift}")
                worksheet0.write(row, 1, avg)
                row += 1
        if full:
            for i in range(self.idx):
                worksheet1.write(i+1, 0, i*self.dt)
            worksheet1.write(0, 0, "time t (s)")
            col = 1
            for zone_id in self.status_number:
                for status in self.status_number[zone_id]:
                    worksheet1.write(0, col, f"({zone_id, *status}")
                    row = 1
                    for num in self.status_number[zone_id][status]:
                        worksheet1.write(row, col, num)
                        row += 1
                    col += 1
        worksheet2.write(0, 0, "unserved number in each zone")
        for i in range(self.city.n):
            for j in range(self.city.n):
                zone_id = i*self.city.n + j
                worksheet2.write(i+1, j, self.unserved_number[zone_id])   
        workbook.close()
        return ni000

    def export_passenger_time(self, full=False):
        dir_path = os.path.dirname(os.path.realpath(__file__))
        path = f"{dir_path}\{self.folder_name}\pax_time" if self.folder_name != None else f"{dir_path}\pax_time"
        path_exist = os.path.exists(f"{path}")
        if not path_exist:
            os.makedirs(f"{path}")
        for info in self.events.queue:
            p = info[1]
            if p.t_start < self.T*1/3:
                continue
            if p.status == 3:
                self.traveling_ts[p.zone.id][p.target_zone.id][p.rs_status].append(p.t_end - p.t_start)
        workbook = xw.Workbook(f"{path}\{self.fleet_m}_pax_time.xlsx")
        worksheet0 = workbook.add_worksheet("passenger total traveling time")
        total_sum, total_sum0, total_sum1 = 0, 0, 0
        total_num, total_num0, total_num1 = 0, 0, 0
        for i in range(self.zone_number):
            for j in range(self.zone_number):
                l0, l1 = len(self.traveling_ts[i][j][0]), len(self.traveling_ts[i][j][1])
                total_sum += (sum(self.traveling_ts[i][j][0]) + sum(self.traveling_ts[i][j][1]))
                total_sum0 += sum(self.traveling_ts[i][j][0])
                total_sum1 += sum(self.traveling_ts[i][j][1])
                total_num += (l0 + l1) 
                total_num0 += l0
                total_num1 += l1 
        avg1 = total_sum/total_num if total_num != 0 else -1
        avg2 = total_sum0/total_num0 if total_num0 != 0 else -1
        avg3 = total_sum1/total_num1 if total_num1 != 0 else -1
        worksheet0.write(0, 0, "average passenger door to door traveling time (hr)")
        worksheet0.write(1, 0, avg1)
        worksheet0.write(3, 0, "average caller door to door traveling time (hr)")
        worksheet0.write(4, 0, avg2)
        worksheet0.write(6, 0, "average seeker door to door traveling time (hr)")
        worksheet0.write(7, 0, avg3)
        P = self.export_P_over_lambda()
        worksheet0.write(8, 0, "P/lambda")
        worksheet0.write(9, 0, P)
        workbook.close()
        return avg1

    def export_unserved_record(self):
        dir_path = os.path.dirname(os.path.realpath(__file__))
        path = f"{dir_path}\{self.folder_name}\info_unserved" if self.folder_name != None else f"{dir_path}\info_unserved"
        path_exist = os.path.exists(f"{path}")
        if not path_exist:
            os.makedirs(f"{path}")
        workbook = xw.Workbook(f"{path}\{self.fleet_m}_info_unserved.xlsx")
        ws = [0 for i in range(self.zone_number)]
        for i in range(self.zone_number):
            ws[i] = workbook.add_worksheet(f"Zone_{i+1}")
            ws[i].write(0, 0, "Incoming time"); ws[i].write(0, 1, "Origin"); ws[i].write(0, 2, "Destination")
            ws[i].write(0, 3, "Direction"); 
            for j in range(self.zone_number):
                ws[i].write(0, 4+j, f"n{i+1}0{j+1}0")
        row = [1 for i in range(self.zone_number)]
        for i in self.unserved_record:
            zone_id = i[1]-1
            ws[zone_id].write(row[zone_id], 0, i[0]); ws[zone_id].write(row[zone_id], 1, i[1]); ws[zone_id].write(row[zone_id], 2, i[2])
            ws[zone_id].write(row[zone_id], 3, i[3]); 
            for j in range(self.zone_number):
                ws[zone_id].write(row[zone_id], 4+j, i[4+j]); 
            row[zone_id] += 1
        ws2 = workbook.add_worksheet("unserved number")
        ws2.write(0, 0, "unserved number in each zone")
        for i in range(self.city.n):
            for j in range(self.city.n):
                zone_id = i*self.city.n + j
                ws2.write(i+1, j, self.unserved_number[zone_id])   
        workbook.close()

    def export_delivery_distance(self):
        dir_path = os.path.dirname(os.path.realpath(__file__))
        path = f"{dir_path}\{self.folder_name}\delivery_distance" if self.folder_name != None else f"{dir_path}\delivery_distance"
        path_exist = os.path.exists(f"{path}")
        if not path_exist:
            os.makedirs(f"{path}")
        wb = xw.Workbook(f"{path}\{self.fleet_m}_delivery_distance.xlsx")
        avg_dist, total_num = [], []
        for veh_id in self.fleet.vehicles:
            veh = self.fleet.vehicles[veh_id]
            dist, num = veh.get_delivery_distance()
            # initialize the summary table
            if avg_dist == []:
                for j in range(self.zone_number):
                    avg_dist.append({})
                    total_num.append({})
                    for i in dist[j]: 
                        avg_dist[j][i] = [0, 0]
                        total_num[j][i] = 0
            for j in range(self.zone_number):
                for i in dist[j]: 
                    if dist[j][i] != None:
                        total_num[j][i] += num[j][i]
                        avg_dist[j][i][0] += dist[j][i]
                        avg_dist[j][i][1] += 1
        for j in range(self.zone_number):
            for i in avg_dist[j]: 
                avg_dist[j][i] = avg_dist[j][i][0]/avg_dist[j][i][1] if avg_dist[j][i][1] != 0 else -1
        
        for j in range(self.zone_number):
            ws = wb.add_worksheet(f"Zone_{j+1}")
            idx1_text = {0:f"{j+1}0{j+1}0", 1:f"{j+1}0{j+1}{j+1}", 2:f"{j+1}0j0", 3:f"{j+1}0{j+1}j", 4:f"{j+1}0jk"} 
            idx2_text = {n:f"Case{n+1}" for n in range(5)}
            for i in avg_dist[j]:
                col, row = i
                ws.write(0, col*3, idx1_text[col])
                ws.write(row+1, col*3, idx2_text[row])
                ws.write(row+1, col*3+1, avg_dist[j][i])
                ws.write(row+1, col*3+2, total_num[j][i])
        wb.close()

    def export_rebalance_m_tp(self):
        dir_path = os.path.dirname(os.path.realpath(__file__))
        path = f"{dir_path}\{self.folder_name}\matrix_tp" if self.folder_name != None else f"{dir_path}\matrix_tp"
        path_exist = os.path.exists(f"{path}")
        if not path_exist:
            os.makedirs(f"{path}")
        wb = xw.Workbook(f"{path}\{self.fleet_m}_matrix_tp.xlsx")
        ws1 = wb.add_worksheet("summary")
        ws1.write(0, 0, f"rebalance number"); ws1.write(0, 1, int((self.T/self.rebalance_t_tp)))
        ws1.write(1, 0, f"rebalance headway"); ws1.write(1, 1, self.rebalance_t_tp)
        ws1.write(2, 0, f"rebalance flow")
        sum_m = np.zeros((self.zone_number, self.zone_number))
        for i in self.rebalance_m_tp: sum_m = np.add(sum_m, i)
        for i in range(self.zone_number):
            for j in range(self.zone_number):
                ws1.write(i+3, j, sum_m[i][j]/self.T)
        ws2 = wb.add_worksheet("zone supply demand")
        for i in range(self.zone_number):
            ws2.write(0, i, f"Zone{i+1}")
        for j in range(len(self.b)):
            for i in range(self.zone_number):
                ws2.write(j+1, i, self.b[j][i])
        wb.close()
        return 

    def export_rebalance_m_bound(self):
        dir_path = os.path.dirname(os.path.realpath(__file__))
        path = f"{dir_path}\{self.folder_name}\matrix_bound" if self.folder_name != None else f"{dir_path}\matrix_bound"
        path_exist = os.path.exists(f"{path}")
        if not path_exist:
            os.makedirs(f"{path}")
        wb = xw.Workbook(f"{path}\{self.fleet_m}_matrix_bound.xlsx")
        ws1 = wb.add_worksheet()
        ws1.write(0, 0, "Rebalancing Rate")
        sum_m = np.zeros((self.zone_number, self.zone_number))
        for i in self.rebalance_m_bound: sum_m = np.add(sum_m, i)
        for i in range(self.zone_number):
            for j in range(self.zone_number):
                ws1.write(i+1, j, sum_m[i][j]/self.T)
        wb.close()
        return 

    def temp_i0ij(self):
        o_xs, o_ys = [], []
        d_xs, d_ys = [], []
        for veh_id in self.fleet.vehicles:
            veh = self.fleet.vehicles[veh_id]
            veh.get_delivery_distance()
            temp1 = veh.temp_i0ij_record_origin[0]
            temp2 = veh.temp_i0ij_record_destination[0]    
            o_xs += temp1[0]
            o_ys += temp1[1]
            d_xs += temp2[0]
            d_ys += temp2[1]
        wb = xw.Workbook(f"Testi0ij_{self.fleet_m}.xlsx")
        ws1 = wb.add_worksheet("origin")
        ws2 = wb.add_worksheet("destination")
        ws1.write(0, 0, "X")
        ws1.write(0, 1, "Y")
        ws2.write(0, 0, "X")
        ws2.write(0, 1, "Y")
        for i in range(len(o_xs)):
            ws1.write(i+1, 0, o_xs[i])
            ws1.write(i+1, 1, o_ys[i])
        for i in range(len(d_xs)):
            ws2.write(i+1, 0, d_xs[i])
            ws2.write(i+1, 1, d_ys[i])
        wb.close()
            