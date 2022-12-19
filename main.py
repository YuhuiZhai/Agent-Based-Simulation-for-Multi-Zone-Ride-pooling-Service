from simulation import Simulation
from city import City
lmd_m = []
rebalance_m = []
prob_matrix = []
for i in range(4):
    lmd_m.append([])
    rebalance_m.append([])
    prob_matrix.append([])
    for j in range(4):
        # if j == 2:
        #     lmd_m[i].append(1400)
        # else:
        #     lmd_m[i].append(200)
        lmd_m[i].append(1000)
        rebalance_m[i].append(0)
        prob_matrix[i].append([0, 0, 0, 0])
prob_matrix[0][1][1] = 1.0
prob_matrix[0][2][3] = 1.0
prob_matrix[0][3][1] = 0.5
prob_matrix[0][3][3] = 0.5
prob_matrix[1][0][2] = 1.0
prob_matrix[1][2][2] = 0.5
prob_matrix[1][2][3] = 0.5
prob_matrix[1][3][3] = 1.0
prob_matrix[2][0][0] = 1.0
prob_matrix[2][1][0] = 0.5
prob_matrix[2][1][1] = 0.5
prob_matrix[2][3][1] = 1.0
prob_matrix[3][0][0] = 0.5
prob_matrix[3][0][2] = 0.5
prob_matrix[3][1][0] = 1.0
prob_matrix[3][2][2] = 1.0
import openpyxl as op
wb = op.load_workbook("test.xlsx")
ws = wb.active
for i in range(2, 17):
    print(f"-------{i-1}-------")
    fleet_m = [[0, 0], [0, 0]]
    fleet_m[0][0] = ws.cell(row=i, column=1).value
    fleet_m[0][1] = ws.cell(row=i, column=2).value
    fleet_m[1][0] = ws.cell(row=i, column=3).value
    fleet_m[1][1] = ws.cell(row=i, column=4).value
    rebalance_m[2][0] = ws.cell(row=i, column=5).value
    rebalance_m[2][1] = ws.cell(row=i, column=6).value
    rebalance_m[2][3] = ws.cell(row=i, column=7).value
    print(fleet_m)
    print(rebalance_m)
    city = City(length=10, n=2, max_v=25)
    s = Simulation(city=city, dt=1/3600, T=10, lmd_matrix=lmd_m, fleet_matrix=fleet_m, rebalance_matrix=rebalance_m)
    s.simple_serve()
    s.export_state_number()
    s.export_passenger_time()
    s.export_P_over_lambda()
    s.export_deliver_dist()
    s.export_uneserved_record()
    s.export_status_duration()



